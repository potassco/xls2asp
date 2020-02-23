#!/usr/bin/env python3
"""
Converts an instance given as a set of excel tables
into a set of asp facts.

Input: Excel xlsx file

Output: Logic program instance file
"""

import warnings
import csv
import argparse
import sys
import traceback
import openpyxl as xls
import math
import warnings
import re
import datetime
from operator import itemgetter

# list all styles and types
list_of_styles = ["sparse_matrix_xy", "matrix_xy", "row", "row_indexed"]
list_of_types = ["auto_detect", "skip", "int",
                 "constant", "time", "date", "datetime", "string"]


def write_category_comment(output, pred):
    output.write('%' * (len(pred) + 6) + '\n')
    output.write('%% ' + pred + ' %%\n')
    output.write('%' * (len(pred) + 6) + '\n')


class Xls2AspError(ValueError):
    def __init__(self, msg, sheet="?", cell=(1, 0)):
        super(Xls2AspError, self).__init__(msg)
        self.sheet = sheet
        self.cell = cell


class TableNameError(ValueError):
    def __init__(self, table):
        ValueError.__init__(
            self, "Name of a tables must respect the syntax of gringo constants", table)


class SheetRowColumnWrongTypeValueError(ValueError):
    def __init__(self, table, row, col, msg, value=None):
        ValueError.__init__(self, 'Wrong type in sheet "{}" row "{}" column "{}": {}'.format(
            table, row, xls.utils.cell.get_column_letter(col+1), msg), value)


class Conversion:

    @staticmethod
    def col2letter(col):
        return xls.utils.cell.get_column_letter(col)

    @staticmethod
    def date2tuple(value):
        return "("+str(value.day)+","+str(value.month)+","+str(value.year)+")"

    @staticmethod
    def datetime2tuple(value):
        return "("+Conversion.date2tuple(value)+","+Conversion.time2tuple(value)+")"

    @staticmethod
    def time2tuple(value):
        return "("+str(value.hour)+","+str(value.minute)+","+str(value.second)+")"

    @staticmethod
    def is_int(value):
        return Conversion.is_single_int(value) or Conversion.is_set_of_int(value)

    @staticmethod
    def is_single_int(value):
        try:
            return int(value) == float(value)
        except (TypeError, ValueError, AttributeError):
            return False

    @staticmethod
    def is_set_of_int(value):
        try:
            for i in value.split(";"):
                a = int(i) == float(i)
            return True
        except (TypeError, ValueError, AttributeError):
            return False
        return False

    @staticmethod
    def normalize_int(value):
        if Conversion.is_single_int(value):
            return value
        else:
            return "("+value+")"

    @staticmethod
    def is_single_string(value):
        s = value.split(";")
        if len(s) >= 2:
            return False
        else:
            return True

    @staticmethod
    def normalize_string(value):
        if Conversion.is_single_string(value):
            return "\""+value+"\""
        else:
            split = value.split(";")
            r = "(" + "\""+split[0]+"\""
            for s in split[1:]:
                r += ";\""+s+"\""
            r += ")"
            return r

    @staticmethod
    def make_predicate(value):
        if Conversion.is_single_constant(value):
            return value
        else:
            val = value[0].lower()+value[1:]
            if Conversion.is_single_constant(val):
                return val
            else:
                raise TableNameError(value)

    @staticmethod
    def is_single_constant(value):
        """
        ensures gringo constant syntax
        """
        const_regex = "_*[a-z][A-Za-z0-9_']*"
        if not isinstance(value, str):
            return False
        m = re.fullmatch(const_regex, value)
        if m != None:
            return True
        else:
            return False

    @staticmethod
    def is_set_of_constant(value):
        if not isinstance(value, str):
            return False
        for s in value.split(";"):
            if not Conversion.is_single_constant(s):
                return False
        return True

    @staticmethod
    def is_asp_constant(value):
        """
        ensures lowercase and no leading or trailing blanks, no whitespace in between
        """
        return Conversion.is_single_constant(value) or Conversion.is_set_of_constant(value)

    @staticmethod
    def normalize_constant(value):
        if Conversion.is_single_constant(value):
            return value
        else:
            return "("+value+")"


class Template:
    """
    Class for reading template
    """

    def __init__(self):
        self.template = {}

    def read(self, fileName):
        with open(fileName, "r") as f:
            for line in f:
                line = line.split("%")[0]
                reader = csv.reader([line], skipinitialspace=True)
                line = next(reader)
                if len(line) != 0:
                    table = line[0]
                    self.add_table(table)
                    style = line[1].strip()
                    if style not in list_of_styles:
                        raise ValueError('style not valid: '+style)
                    self.add_style(table, style)
                    types = line[2:]
                    if style == "matrix_xy":
                        if len(types) != 3:
                            raise ValueError(
                                '3 types are needed to read in matrix style')
                    default = []
                    for t in types:
                        s = t.split("=")
                        if len(s) > 1:
                            default.append(s[1].strip())
                        else:
                            default.append(None)
                        type = s[0].strip()
                        types[types.index(t)] = type
                        if type not in list_of_types:
                            raise ValueError('type not valid: '+t)
                    self.add_types(table, types)
                    self.add_default(table, default)
            f.close()

    def add_table(self, table):
        """
        Adds a table and ensures it is unique
        """
        assert table not in self.template, "Duplicate table '%r' in template" % table
        self.template.setdefault(table, {})

    def add_types(self, table, types):
        """
        Adds a predicate types to a table
        """
        self.template.setdefault(table, {}).setdefault("types", types)

    def add_style(self, table, style):
        self.template.setdefault(table, {}).setdefault("style", style)

    def add_default(self, table, value):
        self.template.setdefault(table, {}).setdefault("default", value)


class Instance:
    """
    Class for maintaining data of an instance file
    """

    def __init__(self, template):
        self.data = {}
        self.template = template

    def add_table(self, table):
        """
        Adds a table and ensures it is unique
        """
        assert table not in self.data, "Duplicate table '%r'" % table
        self.data.setdefault(table, {})

    def correct_table_name(self, table, newname):
        assert newname not in self.data, "Duplicate table '%r' in template" % table
        self.data.setdefault(newname, self.data[table])
        assert newname not in self.template, "Duplicate table '%r' in template" % table
        self.template.setdefault(newname, self.template[table])

    def add_skip(self, table, col=None):
        """
        Adds the index of a column to skip
        """
        if col == None:
            self.data.setdefault(table, {}).setdefault("skip", [])
        else:
            self.data.setdefault(table, {}).setdefault("skip", []).append(col)

    def is_skip(self, table, col):
        try:
            self.data[table]["skip"]
        except (KeyError):
            return False
        return col in self.data[table]["skip"]

    def add_style(self, table, style):
        """
        Adds style to a table
        """
        self.data.setdefault(table, {}).setdefault("style", style)

    def add_row(self, table, id, row):
        self.data.setdefault(table, {}).setdefault(
            "rows", {}).setdefault(id, row)

    def write(self, file):
        for table in self.data:
            style = self.data[table]["style"]
            if style in ["row", "row_indexed"]:
                self.write_table_row_style(table, file, style == 'row_indexed')
            elif style in ["matrix_xy", "sparse_matrix_xy"]:
                self.write_table_matrix_xy_style(
                    table, file, style == 'sparse_matrix_xy')

    def write_table_row_style(self, table, file, prefix_index_argument=False):
        """
        Writes table content to facts row by row
        """
        write_category_comment(file, table)
        for index, row in enumerate(self.data[table]["rows"], 0):
            pred = table+'('
            if prefix_index_argument:
                pred += str(index) + ','
            for col in range(len(self.data[table]["rows"][row])):
                if not self.is_skip(table, col):
                    pred += str(self.data[table]["rows"][row][col])+','
            pred = pred[0:len(pred)-1]
            pred += ').\n'
            file.write(pred)
        file.write('\n')
        file.write('\n')

    def write_table_matrix_xy_style(self, table, file, sparse=False):
        """
        Writes table content to facts
        """
        write_category_comment(file, table)
        for r in self.data[table]["rows"]:
            if r != 1:
                y = self.data[table]["rows"][r][0]
                for col in range(1, len(self.data[table]["rows"][r])):
                    if not self.is_skip(table, col):
                        if not sparse or self.data[table]["rows"][r][col] != None:
                            pred = table + \
                                '('+str(self.data[table]["rows"][1][col])+','
                            pred += str(y)+','
                            pred += str(self.data[table]
                                        ["rows"][r][col])+').\n'
                            file.write(pred)
        file.write('\n')
        file.write('\n')

    def get_test(self, type):
        if type == "int":
            return self.test_int
        elif type == "constant":
            return self.test_constant
        elif type == "string":
            return self.test_string
        elif type == "time":
            return self.test_time
        elif type == "time2time":
            return self.test_time
        elif type == "date":
            return self.test_date
        elif type == "datetime":
            return self.test_datetime
        elif type == "skip":
            return None
        elif type == "auto_detect":
            return self.test_auto_detect
        else:
            raise ValueError('Type not valid: '+type)

    def test_string(self, table, row, col, value, default):
        if value == None and default != None:
            return default
        if isinstance(value, str):
            return Conversion.normalize_string(value)
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting a string, getting:", value)

    def test_int(self, table, row, col, value, default):
        if value == None and default != None:
            return default
        if Conversion.is_int(value):
            return Conversion.normalize_int(value)
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting an int, getting:", value)

    def test_constant(self, table, row, col, value, default=None):
        if value == None and default != None:
            return default
        if Conversion.is_asp_constant(value):
            return Conversion.normalize_constant(value)
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting a constant, getting:", value)

    def test_time(self, table, row, col, value, default):
        if value == None and default != None:
            return default
        if isinstance(value, datetime.time):
            return Conversion.time2tuple(value)
        if value == datetime.datetime(1899, 12, 30, 0, 0):
            print(
                "Warning in table", table, "row ", row, "col ", col)
            print("Expected a time, getting:", value)
            print(
                "This could a know XLS error for times like 00:00:00, treating this as datetime.time(00:00:00).")
            return "(0,0,0)"
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting a time, getting:", value)

    def test_time2min(self, table, row, col, value, default=None):
        if value == None and default != None:
            return default
        if isinstance(value, datetime.time):
            return str(value.hour)+"*60+"+str(value.minute)
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting a time, getting:", value)

    def test_datetime(self, table, row, col, value, default=None):
        if value == None and default != None:
            return default
        if isinstance(value, datetime.datetime):
            return Conversion.datetime2tuple(value)
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting a datetime, getting:", value)

    def test_date(self, table, row, col, value, default=None):
        if value == None and default != None:
            return default
        if isinstance(value, datetime.date):
            return Conversion.date2tuple(value)
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "Expecting a date, getting:", value)

    def test_auto_detect(self, table, row, col, value, default=None):
        if value == None and default != None:
            return default
        if Conversion.is_int(value):
            return Conversion.normalize_int(value)
        elif isinstance(value, datetime.time):
            return Conversion.time2tuple(value)
        elif isinstance(value, datetime.datetime):
            return Conversion.datetime2tuple(value)
        elif isinstance(value, datetime.date):
            return Conversion.date2tuple(value)
        if Conversion.is_asp_constant(value):
            return Conversion.normalize_constant(value)
        elif isinstance(value, str):
            return "\""+value+"\""
        else:
            raise SheetRowColumnWrongTypeValueError(
                table, row, col, "A value is expected", value)

    def correct(self):
        # correct table names
        data = {}
        template = {}
        for table in self.data:
            newname = Conversion.make_predicate(table)
            if newname != table:
                assert newname not in data, "Duplicate table '%r' in template" % table
                data.setdefault(newname, self.data[table])
                assert newname not in self.template, "Duplicate table '%r' in template" % table
                template.setdefault(newname, self.template[table])
        self.data = data
        self.template = template

        # remove leading or trailing blanks from each value in every table
        for table in self.data:
            for r in self.data[table]["rows"]:
                row = self.data[table]["rows"][r]
                for value in row:
                    try:
                        row[row.index(value)] = value.strip()
                    except (AttributeError):
                        pass
        for table in self.data:
            style = self.template[table]["style"]
            if style in ["row", "row_indexed"]:
                self.correct_row_style(table)
            elif style in ["matrix_xy", "sparse_matrix_xy"]:
                self.correct_matrix_xy_style(
                    table, style == "sparse_matrix_xy")
            else:
                raise ValueError('style not valid: '+style)

    def correct_row_style(self, table):
        unexpected = 0
        nb_col = len(self.template[table]["types"])
        self.data[table]["rows"].pop(1)  # ignore first line
        self.ignore_empty_row(table)
        for row in self.data[table]["rows"]:
            if len(self.data[table]["rows"][row]) > nb_col:
                unexpected = 1
                self.data[table]["rows"][row] = self.data[table]["rows"][row][0:nb_col]
        if unexpected:
            sys.stderr.write(
                "WARNING: Undefined column in sheet \""+table+"\", ignoring it\n")
        col = 0
        for i in range(len(self.template[table]["types"])):
            type = self.template[table]["types"][i]
            default = self.template[table]["default"][i]
            if type == "skip":
                self.add_skip(table, col)
            else:
                test = self.get_test(type)
                for row in self.data[table]["rows"]:
                    value = self.data[table]["rows"][row][col]
                    self.data[table]["rows"][row][col] = test(
                        table, row, col, value, default)
            col += 1

    def correct_matrix_xy_style(self, table, sparse=False):
        type_x = self.template[table]["types"][0]
        default_x = self.template[table]["default"][0]
        type_y = self.template[table]["types"][1]
        default_y = self.template[table]["default"][1]
        type_v = self.template[table]["types"][2]
        default_v = self.template[table]["default"][2]

        self.locate_empty_column(table)
        self.add_skip(table, 0)
        self.ignore_empty_row(table)

        # test type for x (= first line)
        test = self.get_test(type_x)
        row_x = self.data[table]["rows"][1]
        # for i in range(1,len(row_x)):
        for col in range(1, len(row_x)):
            if not self.is_skip(table, col):
                row_x[col] = test(table, 1, col, row_x[col], default_x)

        # test type for y (= first column)
        test = self.get_test(type_y)
        for r in self.data[table]["rows"]:
            if r != 1:
                self.data[table]["rows"][r][0] = test(
                    table, r, 0, self.data[table]["rows"][r][0], default_y)

        # test type for the inner matrix
        test = self.get_test(type_v)
        for r in self.data[table]["rows"]:
            if r != 1:
                for col in range(1, len(self.data[table]["rows"][r])):
                    if not self.is_skip(table, col):
                        if not sparse or self.data[table]["rows"][r][col] != None:
                            self.data[table]["rows"][r][col] = test(
                                table, r, col, self.data[table]["rows"][r][col], default_v)

    def get_table_style(self, table):
        if table not in self.template:
            sys.stderr.write("WARNING: Sheet \""+table +
                             "\" is not defined in the template\n")
            return "skip"
        style = self.template[table]["style"]
        return style

    def ignore_empty_row(self, table):
        list_empty = []
        for row in self.data[table]["rows"]:
            empty = 1
            for value in self.data[table]["rows"][row]:
                if value != None:
                    empty = 0
            if empty == 1:
                list_empty.append(row)
        for row in list_empty:
            self.data[table]["rows"].pop(row)
            sys.stderr.write("WARNING: Row "+str(row) +
                             " in sheet \""+table+"\"is empty, ignoring it\n")

    def locate_empty_column(self, table):
        self.add_skip(table)
        for col in range(len(self.data[table]["rows"][1])):
            empty = True
            for row in self.data[table]["rows"]:
                if self.data[table]["rows"][row][col] != None:
                    empty = False
                    break
            if empty:
                self.add_skip(table, col)
        for col in self.data[table]["skip"]:
            sys.stderr.write("WARNING: Column "+Conversion.col2letter(col+1) +
                             " in sheet \""+table+"\" is empty, ignoring it\n")


class XlsReader:

    def __init__(self, instance):
        # Expected worksheets xlsx file and their parsing functions
        self.instance = instance
        self.active_cell = (1, 0)

    def parse(self, input):
        """
        Parses input excel table
        """
        wb = xls.load_workbook(input, read_only=True, data_only=True)
        if self.__update_dimensions(wb):
            wb.close()
            wb = xls.load_workbook(input, read_only=False, data_only=True)
        for sheet in wb:
            style = self.instance.get_table_style(sheet.title)
            if style == "skip":
                sys.stderr.write("Skipping Sheet: "+sheet.title+"\n")
            else:
                self.parse_table(sheet, style)
        for table in self.instance.template:
            if table not in self.instance.data:
                raise ValueError("Sheet \""+table+"\" not found")
            if not self.instance.data[table].__contains__("rows"):
                sys.stderr.write("WARNING: Sheet \""+table +
                                 "\" is empty, ignoring it\n")
                self.instance.data.pop(table)
                sys.stderr.write("Skipping Sheet: "+table+"\n")

    def parse_table(self, sheet, style):
        table = sheet.title
        sys.stderr.write("Parsing Sheet \""+table +
                         "\" with style \""+style+"\"\n")
        self.instance.add_table(sheet.title)
        self.instance.add_style(sheet.title, style)
        self.active_cell = (1, 0)
        self.active_sheet = sheet
        try:
            id = 1
            for r in sheet.iter_rows(min_row=1):
                row = self.parse_row(r)
                self.instance.add_row(table, id, row)
                id += 1
        except Exception as e:
            raise Xls2AspError(str(e), self.active_sheet, self.active_cell)

    def parse_row(self, row, first=0):
        cols = []
        for i in range(first, len(row)):
            cols.append(row[i].value)
        return cols

    def __update_dimensions(self, workbook):
        for sheet in workbook.worksheets:
            if sheet.max_column > 50 or sheet.max_row > 1000:
                return True
        return False


def main():
    # temporal solution, to be removed eventually
    if sys.version_info < (3, 5):
        raise SystemExit('Sorry, this code need Python 3.5 or higher')
    try:
        parser = argparse.ArgumentParser(
            description="Converts an input table to facts"
        )
        parser.add_argument('--output', '-o', metavar='<file>',
                            help='Write output into %(metavar)s', default=sys.stdout, required=False)
        parser.add_argument('--xls', '-x', metavar='<file>',
                            help='Read xls file from %(metavar)s', required=True)
        parser.add_argument('--template', '-t', metavar='<file>',
                            help='Read template from %(metavar)s', required=True)

        args = parser.parse_args()
        tpl = Template()
        tpl.read(args.template)
        instance = Instance(tpl.template)
        reader = XlsReader(instance)
        reader.parse(args.xls)
        instance.correct()
        if args.output == sys.stdout:
            instance.write(args.output)
        else:
            with open(args.output, 'w', encoding="utf8") as f:
                instance.write(f)
        return 0
    except Xls2AspError as e:
        sys.stderr.write("*** Exception: {}\n".format(e))
        sys.stderr.write("***   In sheet={0}:{1}{2}\n".format(
            e.sheet, xls.utils.cell.get_column_letter(e.cell[1]), e.cell[0]))
        return 1
    except Exception as e:
        traceback.print_exception(*sys.exc_info())
        return 1


if __name__ == '__main__':
    sys.exit(main())
